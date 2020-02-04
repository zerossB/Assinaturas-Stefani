import os
import sys

from openpyxl import load_workbook

dir_path = os.path.dirname(os.path.realpath(__file__))


class Workbook(object):
    def __init__(self, workbook=os.path.join('xlsx', 'Assinaturas.xlsx'), sheet="Sheet1"):
        self.workbook = os.path.join(dir_path, workbook)
        self.sheet = sheet

        self.check_files()

    def check_files(self):
        """ simples checagem dos arquivos principais """
        print("*** Checando planilhas ***")
        if not os.path.exists(self.workbook):
            print("Não existe o arquivo xlsx %s" % self.workbook)
            sys.exit(1)
        else:
            print("Checagem concluida.\n\n")
            return True

    def get_rows(self):
        wb = load_workbook(filename=self.workbook)
        sheet = wb[self.sheet]
        return enumerate(sheet)
